from openpyxl import Workbook
from openpyxl import load_workbook

workbook = Workbook() 

sheet = workbook.active #activate excel worksheets


sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx") #create excel docs

work_book = load_workbook(filename="sample.xlsx") #opens the filename and returns the workbook

print(work_book.sheetnames) #prints title as a list 

sheet = work_book.active
print(sheet) #prints name of the worksheet

print(sheet.title) #prints title of the worksheet as a str

print(sheet["A1"]) #prints out a cell

print(sheet["A1"].value) #prints the value of the cell A1

print(sheet.cell(row=10, column=6)) #another way that prints out a cell

print(sheet.cell(row=10, column=6).value) #another value print

